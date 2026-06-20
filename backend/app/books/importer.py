"""Bulk book import: parse CSV / Excel uploads into normalised row dicts.

Kept separate from the route handler so the parsing/validation logic is easy to
test and reason about. The importer only turns an uploaded file into a list of
clean ``{column: value}`` dicts (plus per-row parse errors); the route applies
the database rules (create vs. add copies).
"""
import csv
import io

# Columns the template exposes, in order. ``title``/``author``/``isbn`` are
# required; the rest are optional.
TEMPLATE_COLUMNS = [
    "title",
    "author",
    "isbn",
    "category",
    "publisher",
    "published_year",
    "total_copies",
]

REQUIRED_COLUMNS = ("title", "author", "isbn")

# A filled-in example row shown in the downloadable template.
EXAMPLE_ROW = {
    "title": "Clean Code",
    "author": "Robert C. Martin",
    "isbn": "9780132350884",
    "category": "Software",
    "publisher": "Prentice Hall",
    "published_year": "2008",
    "total_copies": "3",
}


class ImportError_(Exception):
    """Raised when the uploaded file cannot be parsed at all."""


def _normalise_header(name):
    return (name or "").strip().lower().replace(" ", "_")


def parse_upload(filename, raw_bytes):
    """Parse an uploaded CSV/XLSX file into ``(rows, headers)``.

    ``rows`` is a list of ``{column: str_value}`` dicts keyed by the normalised
    template columns. Raises :class:`ImportError_` for unreadable files or a
    missing required column.
    """
    name = (filename or "").lower()
    if name.endswith(".csv") or name.endswith(".txt"):
        rows = _parse_csv(raw_bytes)
    elif name.endswith(".xlsx"):
        rows = _parse_xlsx(raw_bytes)
    else:
        raise ImportError_(
            "Unsupported file type. Upload a .csv or .xlsx file."
        )

    if not rows:
        raise ImportError_("The file has no data rows.")

    headers = set(rows[0].keys())
    missing = [c for c in REQUIRED_COLUMNS if c not in headers]
    if missing:
        raise ImportError_(
            "Missing required column(s): " + ", ".join(missing)
        )
    return rows


def _parse_csv(raw_bytes):
    text = raw_bytes.decode("utf-8-sig", errors="replace")
    reader = csv.DictReader(io.StringIO(text))
    if reader.fieldnames is None:
        raise ImportError_("The CSV file is empty.")
    field_map = {f: _normalise_header(f) for f in reader.fieldnames}
    rows = []
    for raw in reader:
        row = {field_map[k]: (v or "").strip() for k, v in raw.items() if k in field_map}
        if any(row.values()):  # skip fully blank lines
            rows.append(row)
    return rows


def _parse_xlsx(raw_bytes):
    # Imported lazily so the rest of the app works even if openpyxl is absent.
    from openpyxl import load_workbook

    try:
        wb = load_workbook(io.BytesIO(raw_bytes), read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001 - surface a friendly message
        raise ImportError_(f"Could not read the Excel file: {exc}")
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = next(rows_iter)
    except StopIteration:
        raise ImportError_("The Excel file is empty.")

    headers = [_normalise_header(str(h)) if h is not None else "" for h in header]
    rows = []
    for values in rows_iter:
        row = {}
        for idx, header_name in enumerate(headers):
            if not header_name:
                continue
            value = values[idx] if idx < len(values) else None
            row[header_name] = "" if value is None else str(value).strip()
        if any(row.values()):
            rows.append(row)
    return rows
